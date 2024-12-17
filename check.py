from openpyxl import load_workbook

# 加载 Excel 文件
file_path = r'C:\Users\user\Desktop\SuChenAI\Factbook-RPA-clickurl\data\fb_account_data.xlsx'
wb = load_workbook(file_path)

# 打印工作表名称以确认工作簿是否正确加载
print("工作簿中的工作表:", wb.sheetnames)

# 访问 '跳轉' 工作表
try:
    sheet = wb["跳轉"]
    print(f"成功加载工作表: {sheet.title}")
    # 打印工作表的所有数据
    for row in sheet.iter_rows(values_only=True):
        print(row)
except KeyError:
    print("工作表 '跳轉' 不存在，请检查文件或名称拼写是否正确。")